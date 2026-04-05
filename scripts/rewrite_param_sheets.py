"""
Reescribe las hojas ANALISIS_Y_DEFINICIONES_COTIZAD y CALCULOS_FUJO en INPUT_FILES.xlsx
como tablas de parámetros consumibles por el programa.

Nuevas hojas:
  REGLAS_INMOBILIARIAS  — tipo de cálculo y LTV por alianza
  PARAMETROS_CALCULO    — constantes del motor de cálculo
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook('INPUT_FILES.xlsx')

# Eliminar hojas antiguas si existen
for name in ['ANALISIS_Y_DEFINICIONES_COTIZAD', 'CALCULOS_FUJO',
             'REGLAS_INMOBILIARIAS', 'PARAMETROS_CALCULO']:
    if name in wb.sheetnames:
        del wb[name]

thin   = Side(border_style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap   = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

def hfill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def apply(cell, font=None, fill=None, align=None):
    if font:  cell.font      = font
    if fill:  cell.fill      = fill
    if align: cell.alignment = align
    cell.border = border

H_FONT  = Font(bold=True, color='FFFFFF', size=11)
H_FILL  = hfill('1F4E79')
B_FONT  = Font(bold=True, size=10)
N_FONT  = Font(size=10)
Y_FILL  = hfill('FFE699')
G_FILL  = hfill('D9EAD3')
B_FILL  = hfill('D0E4F7')
W_FILL  = hfill('FFFFFF')
S_FILL  = hfill('E2EFDA')

# ═══════════════════════════════════════════════════════════════
# HOJA 1: REGLAS_INMOBILIARIAS
# ═══════════════════════════════════════════════════════════════
ws1 = wb.create_sheet('REGLAS_INMOBILIARIAS')

headers1 = [
    'ALIANZA',
    'TIPO_CALCULO_BONO',
    'LTV_MAX_PCT',
    'PIE_CONJUNTOS_PCT',
    'DESCRIPCION_BONO_PIE',
]
for col, h in enumerate(headers1, 1):
    c = ws1.cell(row=1, column=col, value=h)
    apply(c, font=H_FONT, fill=H_FILL, align=center)

# Datos: una fila por inmobiliaria conocida
# TIPO_CALCULO_BONO valores válidos: maestra | precio-lista-depto | precio-lista-total
reglas = [
    ('MAESTRA',   'maestra',            0.80, 0.20,
     'Bono sobre TASACION BANCO (fórmula D35/D36 Excel). Crédito hipotecario = tasación × LTV_MAX_PCT.'),
    ('INGEVEC',   'precio-lista-depto', 1.00, 0.20,
     'Bono = Precio Lista DEPTO × %BonoPie. Tasación = Valor Venta + Bono. Crédito = Valor Venta - Pie - Aporte.'),
    ('RVC',       'precio-lista-depto', 1.00, 0.20,
     'Bono = Precio Lista DEPTO × %BonoPie. Tasación = Valor Venta + Bono. Crédito = Valor Venta - Pie - Aporte.'),
    ('URMENETA',  'precio-lista-total', 1.00, 0.20,
     'Bono = Precio Lista TOTAL (depto+estac+bodega) × %BonoPie. Tasación = Valor Venta + Bono. Crédito = Valor Venta - Pie - Aporte.'),
    ('TOCTOC',    'precio-lista-depto', 1.00, 0.20,
     'Bono = Precio Lista DEPTO × %BonoPie. Tasación = Valor Venta + Bono. Crédito = Valor Venta - Pie - Aporte.'),
    ('EURO',      'precio-lista-depto', 1.00, 0.20,
     'Bono = Precio Lista DEPTO × %BonoPie. Tasación = Valor Venta + Bono. Crédito = Valor Venta - Pie - Aporte.'),
]

fill_map = {
    'MAESTRA':  Y_FILL,
    'URMENETA': B_FILL,
}

for row_idx, (alianza, tipo, ltv, pie_conj, desc) in enumerate(reglas, 2):
    fill = fill_map.get(alianza, G_FILL)
    values = [alianza, tipo, ltv, pie_conj, desc]
    for col, val in enumerate(values, 1):
        c = ws1.cell(row=row_idx, column=col, value=val)
        apply(c, font=B_FONT if col == 1 else N_FONT, fill=fill, align=wrap)

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 70
ws1.row_dimensions[1].height = 20
for r in range(2, len(reglas)+2):
    ws1.row_dimensions[r].height = 45
ws1.freeze_panes = 'A2'

# ═══════════════════════════════════════════════════════════════
# HOJA 2: PARAMETROS_CALCULO
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('PARAMETROS_CALCULO')

headers2 = ['PARAMETRO', 'VALOR', 'TIPO', 'DESCRIPCION']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    apply(c, font=H_FONT, fill=H_FILL, align=center)

# TIPO: number | decimal | percent
parametros = [
    # Constantes del modelo de inversión
    ('MESES_ARRIENDO_ANIO',    11,    'number',
     'Meses de arriendo por año asumidos (1 mes vacío). Usado en Cap Rate y flujo acumulado.'),
    ('HAIRCUT_VENTA',          0.95,  'decimal',
     'Factor de castigo sobre valor de venta proyectado al año 5 (simula costos de transacción).'),
    ('PLUSVALIA_ANUAL_DEFAULT', 0.02,  'decimal',
     'Plusvalía anual estimada por defecto (2%). Editable en UI.'),

    # Parámetros de pie
    ('PIE_CONJUNTOS_PCT',      0.20,  'decimal',
     'Porcentaje de pie obligatorio para bienes conjuntos (estacionamiento, bodega). Fijo, no editable.'),
    ('UPFRONT_PCT_DEFAULT',    0.02,  'decimal',
     'Upfront a la Promesa por defecto (2%). Editable en UI. Aplica a todas las inmobiliarias.'),
    ('PIE_DEFAULT_PCT',        0.10,  'decimal',
     'Porcentaje de pie por defecto para el departamento cuando no hay bono pie.'),
    ('PIE_CON_BONO_FORMULA',   '0.20 - BONO_PIE_PCT', 'formula',
     'Fórmula para calcular pie default cuando hay bono pie: max(0, 0.20 - %BonoPie).'),

    # Parámetros hipotecarios
    ('PLAZO_DEFAULT_ANIOS',    30,    'number',
     'Plazo hipotecario por defecto en años.'),
    ('CAE_DEFAULT_1',          0.040, 'decimal',
     'Tasa CAE por defecto escenario 1.'),
    ('CAE_DEFAULT_2',          0.045, 'decimal',
     'Tasa CAE por defecto escenario 2.'),
    ('CAE_DEFAULT_3',          0.050, 'decimal',
     'Tasa CAE por defecto escenario 3.'),

    # Fórmulas de evaluación (referencia)
    ('FORMULA_CAP_RATE',       '(arriendo_mensual * MESES_ARRIENDO_ANIO / valor_UF) / tasacion_UF', 'formula',
     'Cap Rate anual: renta bruta anual en UF sobre tasación banco.'),
    ('FORMULA_ROI_5_ANIOS',    '(precio_venta_anio5 - valor_venta + flujo_acumulado) / equity_base', 'formula',
     'ROI sobre equity pagado (pie) en 5 años. Incluye plusvalía y flujo de arriendo neto.'),
    ('FORMULA_ROI_ANUAL',      '(1 + ROI_5_ANIOS)^(1/5) - 1', 'formula',
     'ROI anual compuesto equivalente al ROI en 5 años.'),
    ('FORMULA_PRECIO_VENTA_5', 'valor_venta * (1 + plusvalia_anual)^5 * HAIRCUT_VENTA', 'formula',
     'Precio de venta proyectado al año 5 con castigo por costos de transacción.'),
]

for row_idx, (param, val, tipo, desc) in enumerate(parametros, 2):
    fill = S_FILL if tipo == 'formula' else W_FILL
    values = [param, val, tipo, desc]
    for col, v in enumerate(values, 1):
        c = ws2.cell(row=row_idx, column=col, value=v)
        apply(c, font=B_FONT if col == 1 else N_FONT, fill=fill, align=wrap)

ws2.column_dimensions['A'].width = 26
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 72
ws2.row_dimensions[1].height = 20
for r in range(2, len(parametros)+2):
    ws2.row_dimensions[r].height = 40
ws2.freeze_panes = 'A2'

wb.save('INPUT_FILES.xlsx')
print('OK — hojas REGLAS_INMOBILIARIAS y PARAMETROS_CALCULO creadas en INPUT_FILES.xlsx')
