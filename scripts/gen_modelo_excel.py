"""
Genera MODELO_DATOS_COTIZADOR.xlsx
Documenta tablas, columnas, PKs, FKs, UKs, índices, constraints y relaciones.
"""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

# ── Paleta ────────────────────────────────────────────────────────────────────
C_TITLE_BG   = "1F3864"   # azul muy oscuro  — encabezados de hoja
C_TITLE_FG   = "FFFFFF"
C_TABLE_BG   = "2E75B6"   # azul medio       — nombre de tabla
C_TABLE_FG   = "FFFFFF"
C_PK_BG      = "FFF2CC"   # amarillo claro   — columna PK
C_FK_BG      = "DDEBF7"   # azul claro       — columna FK
C_UK_BG      = "E2EFDA"   # verde claro      — columna UK
C_IDX_BG     = "FCE4D6"   # naranja claro    — índices
C_REL_BG     = "EAD1DC"   # rosa claro       — relaciones
C_CHK_BG     = "F4CCFF"   # violeta claro    — constraints CHECK
C_ODD        = "F2F2F2"   # gris muy claro   — filas impares
C_EVEN       = "FFFFFF"   # blanco           — filas pares
C_SECTION_BG = "D6E4F0"   # celeste          — separador de sección

# Colores únicos por tabla
TABLE_COLORS = {
    "inmobiliaria":        ("BDD7EE", "1F3864"),
    "proyecto":            ("C6EFCE", "375623"),
    "unidad":              ("FFEB9C", "9C5700"),
    "condicion_comercial": ("FCE4D6", "833C00"),
    "bien_conjunto":       ("E2EFDA", "375623"),
    "uf_valor":            ("EAD1DC", "4A235A"),
    "parametro_cotizador": ("D9D2E9", "20124D"),
    "cotizacion":          ("CFE2F3", "1C4587"),
}

# ── Estilos ───────────────────────────────────────────────────────────────────
thin   = Side(style="thin",   color="BBBBBB")
medium = Side(style="medium", color="888888")
thick  = Side(style="medium", color="444444")

def brd(t="thin"):
    s = thin if t == "thin" else medium
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def fmt_header(ws, row, col, value, bg=C_TITLE_BG, fg=C_TITLE_FG,
               size=10, bold=True, wrap=True, halign="center"):
    c = ws.cell(row, col, value)
    c.font      = Font(name="Calibri", bold=bold, color=fg, size=size)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal=halign, vertical="center",
                             wrap_text=wrap)
    c.border    = Border(left=thick, right=thick, top=thick, bottom=thick)
    return c

def fmt_cell(ws, row, col, value, bg=C_EVEN, bold=False,
             size=9, halign="left", wrap=True, italic=False):
    c = ws.cell(row, col, value)
    c.font      = Font(name="Calibri", bold=bold, color="222222",
                       size=size, italic=italic)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal=halign, vertical="top",
                             wrap_text=wrap)
    c.border    = brd("thin")
    return c

def fmt_table_label(ws, row, col_start, col_end, table_name):
    bg, fg = TABLE_COLORS.get(table_name, (C_TABLE_BG, C_TABLE_FG))
    c = ws.cell(row, col_start, f"  {table_name.upper()}")
    c.font      = Font(name="Calibri", bold=True, color=fg, size=10)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = Border(left=thick, right=thick, top=thick, bottom=thick)
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    ws.row_dimensions[row].height = 18

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

# ── Modelo de datos ───────────────────────────────────────────────────────────
TABLES = [
    {
        "name": "inmobiliaria",
        "desc": "Maestro de inmobiliarias (alianzas). Fuente: columna ALIANZA.",
        "rows": 5,
        "fuente": "Manual / columna ALIANZA de PROYECTOS",
        "mantenedor": "Admin — CRUD + baja lógica",
        "columns": [
            # (nombre, tipo, rol, nullable, default, descripcion)
            ("id_inmobiliaria", "INTEGER", "PK",  "NO",  "autoincrement", "Clave primaria surrogate"),
            ("nombre",          "TEXT",    "UK",  "NO",  "—",             "Nombre único: INGEVEC, MAESTRA, RVC, TOCTOC, URMENETA"),
            ("activo",          "INTEGER", "—",   "NO",  "1",             "Baja lógica: 0=inactiva  1=activa"),
            ("created_at",      "TEXT",    "—",   "NO",  "datetime('now')", "Timestamp de creación ISO 8601"),
            ("updated_at",      "TEXT",    "—",   "NO",  "datetime('now')", "Timestamp de última modificación — actualizado por trigger"),
        ],
    },
    {
        "name": "proyecto",
        "desc": "Maestro de proyectos inmobiliarios. Fuente: hoja PROYECTOS (99 filas).",
        "rows": 99,
        "fuente": "Hoja PROYECTOS del Excel",
        "mantenedor": "Admin — CRUD completo",
        "columns": [
            ("id_proyecto",     "INTEGER", "PK",  "NO",  "autoincrement", "Clave primaria surrogate"),
            ("id_inmobiliaria", "INTEGER", "FK",  "NO",  "—",             "FK → inmobiliaria.id_inmobiliaria  ON UPDATE CASCADE  ON DELETE RESTRICT"),
            ("nemotecnico",     "TEXT",    "UK",  "NO",  "—",             "Clave natural única: ABD, TOC, BEL… Usada como referencia en toda la app"),
            ("nombre_proyecto", "TEXT",    "UK²", "NO",  "—",             "UK compuesta con id_inmobiliaria"),
            ("comuna",          "TEXT",    "—",   "NO",  "—",             ""),
            ("direccion",       "TEXT",    "—",   "NO",  "—",             ""),
            ("tipo_entrega",    "TEXT",    "—",   "NO",  "'Entrega Futura'", "CHECK: 'Entrega Inmediata' | 'Entrega Futura'"),
            ("periodo_entrega", "TEXT",    "—",   "NO",  "—",             "Ej: 'Inmediata', '1er semestre 2028'"),
            ("activo",          "INTEGER", "—",   "NO",  "1",             "Baja lógica. Incorpora ESTADO_PROYECTO de CONDICIONES_COMERCIALES"),
            ("created_at",      "TEXT",    "—",   "NO",  "datetime('now')", ""),
            ("updated_at",      "TEXT",    "—",   "NO",  "datetime('now')", "Actualizado por trigger"),
        ],
    },
    {
        "name": "unidad",
        "desc": "Stock de unidades. Fuente: hoja STOCK NUEVOS (8.646 filas, 21 cols).",
        "rows": 8646,
        "fuente": "Hoja STOCK NUEVOS del Excel",
        "mantenedor": "Admin — CRUD + importación masiva desde Excel",
        "columns": [
            ("id_unidad",             "INTEGER", "PK",  "NO",  "autoincrement", "Surrogate key obligatoria: (nemo+tipo+num) tiene 73 duplicados y 916 NULLs"),
            ("id_proyecto",           "INTEGER", "FK",  "NO",  "—",             "FK → proyecto.id_proyecto  ON UPDATE CASCADE  ON DELETE RESTRICT"),
            ("numero_unidad",         "INTEGER", "—",   "SÍ",  "NULL",          "NULL para bodegas/estac sin número asignado (916 casos)"),
            ("tipo_unidad",           "TEXT",    "—",   "NO",  "—",             "CHECK: Departamento | Estacionamiento | Bodega | Local Comercial | Local comercial | Local | Estacionamiento Moto | Estacionamiento Comercial | Estacionamiento local"),
            ("programa",              "TEXT",    "—",   "NO",  "—",             "Ej: '2D1B', '1D1B', 'Bodega', 'Estacionamiento'"),
            ("piso_producto",         "INTEGER", "—",   "NO",  "—",             "Incluye negativos (-4 a -1 = subterráneo)"),
            ("orientacion",           "TEXT",    "—",   "SÍ",  "NULL",          "NORTE | SUR | ORIENTE | PONIENTE | NOR ORIENTE…"),
            ("dormitorios",           "TEXT",    "—",   "SÍ",  "NULL",          "Valores: '1', '2', '3', '1-1/2', 'BO', '#N/A'"),
            ("banios",                "INTEGER", "—",   "SÍ",  "NULL",          "1 o 2"),
            ("superficie_terreno_m2", "REAL",    "—",   "NO",  "0",             "Casas; departamentos/estac/bodegas se importan con 0"),
            ("superficie_util_m2",    "REAL",    "—",   "SÍ",  "NULL",          "ATENCIÓN: origen texto con coma '43,35' → parsear a REAL reemplazando ',' por '.'"),
            ("superficie_terraza_m2", "REAL",    "—",   "SÍ",  "NULL",          "Mismo tratamiento que superficie_util_m2"),
            ("superficie_total_m2",   "REAL",    "—",   "NO",  "—",             "CHECK: >= 0"),
            ("precio_lista_uf",       "REAL",    "—",   "NO",  "—",             "CHECK: > 0  —  valor en UF"),
            ("estado_stock",          "TEXT",    "—",   "NO",  "'Disponible'",  "CHECK: Disponible | Arrendado | En Recolocación | Reservado. NORMALIZAR 'DISPONIBLE'→'Disponible' en importación"),
            ("bienes_conjuntos",      "TEXT",    "—",   "SÍ",  "NULL",          "Texto crudo: 'B - 1', 'B - 1 B'… Se resuelve en tabla bien_conjunto"),
            ("created_at",            "TEXT",    "—",   "NO",  "datetime('now')", ""),
            ("updated_at",            "TEXT",    "—",   "NO",  "datetime('now')", "Actualizado por trigger"),
        ],
    },
    {
        "name": "condicion_comercial",
        "desc": "Condiciones comerciales por proyecto, tipo de unidad y programa. Fuente: hoja CONDICIONES_COMERCIALES (309 filas, 0 duplicados).",
        "rows": 309,
        "fuente": "Hoja CONDICIONES_COMERCIALES del Excel",
        "mantenedor": "Admin — CRUD + importación masiva. Cambios afectan todos los cálculos",
        "columns": [
            ("id_condicion",             "INTEGER", "PK",  "NO",  "autoincrement", "Surrogate key"),
            ("id_proyecto",              "INTEGER", "FK",  "NO",  "—",             "FK → proyecto.id_proyecto  ON UPDATE CASCADE  ON DELETE RESTRICT"),
            ("tipo_unidad",              "TEXT",    "UK",  "NO",  "—",             "UK compuesta: (id_proyecto, tipo_unidad, programa)"),
            ("programa",                 "TEXT",    "UK",  "NO",  "—",             "UK compuesta: (id_proyecto, tipo_unidad, programa)"),
            ("reserva_clp",              "INTEGER", "—",   "NO",  "100000",        "Monto de reserva en CLP. CHECK: >= 0"),
            ("descuento",                "REAL",    "—",   "NO",  "0",             "Porcentaje decimal [0,1]. Ej: 0.10 = 10%. CHECK: BETWEEN 0 AND 1"),
            ("bono_pie",                 "REAL",    "—",   "NO",  "0",             "Porcentaje decimal [0,1]. Actúa como Aporte Inmobiliaria. CHECK: BETWEEN 0 AND 1"),
            ("cuotas_pie",               "INTEGER", "—",   "NO",  "0",             "N° de cuotas del saldo de pie. CHECK: >= 0"),
            ("pie_periodo_construccion", "REAL",    "—",   "NO",  "0",             "% pie para proyectos Entrega Futura. CHECK: BETWEEN 0 AND 1"),
            ("cuoton",                   "REAL",    "—",   "NO",  "0",             "% cuotón especial. CHECK: BETWEEN 0 AND 1"),
            ("pie_credito_directo",      "REAL",    "—",   "NO",  "0",             "% para modalidad crédito directo. CHECK: BETWEEN 0 AND 1"),
            ("activo",                   "INTEGER", "—",   "NO",  "1",             "Baja lógica. CHECK: IN (0,1)"),
            ("created_at",               "TEXT",    "—",   "NO",  "datetime('now')", ""),
            ("updated_at",               "TEXT",    "—",   "NO",  "datetime('now')", "Actualizado por trigger"),
        ],
    },
    {
        "name": "bien_conjunto",
        "desc": "Tabla junction N:M: relaciona Departamentos con sus Estacionamientos y/o Bodegas. Fuente: columna BIENES_CONJUNTOS de STOCK NUEVOS (~409 filas).",
        "rows": 409,
        "fuente": "Derivada de columna BIENES_CONJUNTOS en STOCK NUEVOS",
        "mantenedor": "Admin — CRUD + auto-generación en importación de stock",
        "columns": [
            ("id_bien_conjunto",    "INTEGER", "PK",  "NO",  "autoincrement", "Surrogate key"),
            ("id_unidad_principal", "INTEGER", "FK",  "NO",  "—",             "FK → unidad.id_unidad (siempre un Departamento). ON DELETE CASCADE"),
            ("id_unidad_asociada",  "INTEGER", "FK",  "NO",  "—",             "FK → unidad.id_unidad (siempre Estacionamiento o Bodega). ON DELETE CASCADE"),
            ("descripcion",         "TEXT",    "—",   "SÍ",  "NULL",          "Texto original del Excel para trazabilidad: 'B - 1', 'B - 1 B'…"),
            ("created_at",          "TEXT",    "—",   "NO",  "datetime('now')", ""),
        ],
    },
    {
        "name": "uf_valor",
        "desc": "Tabla de valores diarios de la UF desde 1977-08-01. Fuente: hoja UF (17.784 filas).",
        "rows": 17784,
        "fuente": "Hoja UF del Excel / API CMF",
        "mantenedor": "Admin — INSERT periódico diario + importación masiva anual",
        "columns": [
            ("fecha",            "TEXT", "PK",  "NO",  "—",             "Clave natural 'YYYY-MM-DD' — único por día. Columnas MONEDA y PERIODO del Excel se descartan"),
            ("valor_uf",         "REAL", "—",   "NO",  "—",             "Valor en pesos CLP. CHECK: > 0"),
            ("variacion_diaria", "REAL", "—",   "SÍ",  "NULL",          "Diferencia respecto al día anterior"),
            ("created_at",       "TEXT", "—",   "NO",  "datetime('now')", ""),
        ],
    },
    {
        "name": "parametro_cotizador",
        "desc": "Configuración de la UI y constantes del motor de cálculo. Reemplaza la hoja aux del Excel.",
        "rows": 22,
        "fuente": "Hoja aux del Excel + constantes del COTIZADOR",
        "mantenedor": "Admin senior — CRUD completo. Modificar constantes afecta todos los cálculos",
        "columns": [
            ("id_parametro",   "INTEGER", "PK",  "NO",  "autoincrement", "Surrogate key"),
            ("categoria",      "TEXT",    "UK",  "NO",  "—",             "UK compuesta (categoria, clave). CHECK: CAE | PIE_PCT | PLAZO | CONSTANTE"),
            ("clave",          "TEXT",    "UK",  "NO",  "—",             "UK compuesta (categoria, clave). Identificador único dentro de categoría"),
            ("valor_numerico", "REAL",    "—",   "SÍ",  "NULL",          "Valor usado en cálculos"),
            ("etiqueta",       "TEXT",    "—",   "NO",  "—",             "Texto para mostrar en UI: '4.0%', '30 años', etc."),
            ("orden",          "INTEGER", "—",   "NO",  "0",             "Posición en dropdown de la UI"),
            ("es_default",     "INTEGER", "—",   "NO",  "0",             "1 = valor preseleccionado. CHECK: IN (0,1)"),
            ("activo",         "INTEGER", "—",   "NO",  "1",             "0 = oculto sin eliminar. CHECK: IN (0,1)"),
            ("created_at",     "TEXT",    "—",   "NO",  "datetime('now')", ""),
            ("updated_at",     "TEXT",    "—",   "NO",  "datetime('now')", "Actualizado por trigger"),
        ],
    },
    {
        "name": "cotizacion",
        "desc": "Historial inmutable de cotizaciones generadas [FASE 2]. Snapshot completo para trazabilidad y reimpresión.",
        "rows": 0,
        "fuente": "Generada por la aplicación al crear cada cotización",
        "mantenedor": "Solo lectura para Admin — INSERT desde la app",
        "columns": [
            ("id_cotizacion",       "INTEGER", "PK",  "NO",  "autoincrement", "Surrogate key"),
            ("fecha_generacion",    "TEXT",    "—",   "NO",  "datetime('now')", "Timestamp ISO 8601"),
            ("id_unidad",           "INTEGER", "FK",  "NO",  "—",             "FK → unidad.id_unidad  ON DELETE RESTRICT"),
            ("id_condicion",        "INTEGER", "FK",  "NO",  "—",             "FK → condicion_comercial.id_condicion  ON DELETE RESTRICT"),
            ("nombre_broker",       "TEXT",    "—",   "NO",  "—",             ""),
            ("nombre_cliente",      "TEXT",    "—",   "NO",  "—",             ""),
            ("rut_cliente",         "TEXT",    "—",   "SÍ",  "NULL",          ""),
            ("email_cliente",       "TEXT",    "—",   "SÍ",  "NULL",          ""),
            ("celular_cliente",     "TEXT",    "—",   "SÍ",  "NULL",          ""),
            ("valor_uf_snapshot",   "REAL",    "—",   "NO",  "—",             "UF del día — snapshot para reimpresión futura"),
            ("pie_pct",             "REAL",    "—",   "NO",  "—",             "Porcentaje de pie usado en la cotización"),
            ("n_cuotas_pie",        "INTEGER", "—",   "NO",  "—",             ""),
            ("cae_1",               "REAL",    "—",   "NO",  "—",             "Escenario 1 — ej: 0.040"),
            ("cae_2",               "REAL",    "—",   "NO",  "—",             "Escenario 2 — ej: 0.045"),
            ("cae_3",               "REAL",    "—",   "NO",  "—",             "Escenario 3 — ej: 0.050"),
            ("plazo_anios",         "INTEGER", "—",   "NO",  "—",             "20 | 25 | 30"),
            ("arriendo_1_clp",      "INTEGER", "—",   "SÍ",  "NULL",          ""),
            ("arriendo_2_clp",      "INTEGER", "—",   "SÍ",  "NULL",          ""),
            ("arriendo_3_clp",      "INTEGER", "—",   "SÍ",  "NULL",          ""),
            ("plusvalia_anual_pct", "REAL",    "—",   "NO",  "0.02",          "Plusvalía anual estimada usada en la evaluación"),
            ("precio_lista_uf",     "REAL",    "—",   "NO",  "—",             "Snapshot del precio al momento de cotizar"),
            ("precio_venta_uf",     "REAL",    "—",   "NO",  "—",             "Precio con descuento aplicado"),
            ("tasacion_uf",         "REAL",    "—",   "SÍ",  "NULL",          "NULL cuando bono_pie = 0"),
            ("pie_total_uf",        "REAL",    "—",   "NO",  "—",             ""),
            ("ch_plan_uf",          "REAL",    "—",   "NO",  "—",             "Base del PMT para la simulación hipotecaria"),
            ("roi_5anios",          "REAL",    "—",   "SÍ",  "NULL",          ""),
            ("roi_anual",           "REAL",    "—",   "SÍ",  "NULL",          ""),
            ("cap_rate",            "REAL",    "—",   "SÍ",  "NULL",          ""),
        ],
    },
]

RELATIONS = [
    # (tabla_origen, col_origen, tabla_destino, col_destino, tipo, on_update, on_delete, cardinalidad, descripcion)
    ("proyecto",            "id_inmobiliaria", "inmobiliaria",        "id_inmobiliaria", "FK física",  "CASCADE",  "RESTRICT", "N:1",  "Muchos proyectos pertenecen a una inmobiliaria"),
    ("unidad",              "id_proyecto",     "proyecto",            "id_proyecto",     "FK física",  "CASCADE",  "RESTRICT", "N:1",  "Muchas unidades pertenecen a un proyecto"),
    ("condicion_comercial", "id_proyecto",     "proyecto",            "id_proyecto",     "FK física",  "CASCADE",  "RESTRICT", "N:1",  "Muchas condiciones definen un proyecto"),
    ("bien_conjunto",       "id_unidad_principal", "unidad",          "id_unidad",       "FK física",  "—",        "CASCADE",  "N:1",  "El departamento principal. Si se elimina la unidad, se eliminan sus bienes"),
    ("bien_conjunto",       "id_unidad_asociada",  "unidad",          "id_unidad",       "FK física",  "—",        "CASCADE",  "N:1",  "El estacionamiento/bodega asociado. Misma política CASCADE"),
    ("cotizacion",          "id_unidad",       "unidad",              "id_unidad",       "FK física",  "—",        "RESTRICT", "N:1",  "Una cotización referencia una unidad. No se puede eliminar unidad con cotizaciones"),
    ("cotizacion",          "id_condicion",    "condicion_comercial", "id_condicion",    "FK física",  "—",        "RESTRICT", "N:1",  "Una cotización referencia una condición. No se puede eliminar la condición"),
    ("unidad",              "(id_proyecto, tipo_unidad, programa)", "condicion_comercial", "(id_proyecto, tipo_unidad, programa)", "JOIN lógico", "—", "—", "N:1", "No es FK física. El JOIN se realiza por estos 3 campos en tiempo de ejecución. Una condición aplica a N unidades del mismo programa"),
]

INDEXES = [
    # (tabla, nombre, columnas, tipo, proposito)
    ("proyecto",            "idx_proyecto_inmobiliaria", "id_inmobiliaria",                                   "Normal",         "Lookup FK"),
    ("proyecto",            "idx_proyecto_activo",       "activo, id_inmobiliaria",                           "Normal",         "Filtro de proyectos activos por inmobiliaria"),
    ("proyecto",            "idx_proyecto_cascada",      "comuna, tipo_entrega, id_inmobiliaria, activo  WHERE activo=1", "Parcial", "Dropdown en cascada: Comuna → Entrega → Inmobiliaria → Proyecto"),
    ("unidad",              "uq_unidad_por_proyecto",    "id_proyecto, tipo_unidad, numero_unidad",           "Único PARCIAL WHERE numero_unidad IS NOT NULL", "Integridad: no duplicados reales. Permite múltiples NULLs"),
    ("unidad",              "idx_unidad_proyecto",       "id_proyecto",                                       "Normal",         "Lookup FK"),
    ("unidad",              "idx_unidad_disponibles",    "id_proyecto, estado_stock, tipo_unidad",            "Normal",         "Filtro principal del cotizador (unidades disponibles)"),
    ("unidad",              "idx_unidad_numero",         "numero_unidad",                                     "Normal",         "Selección por número de unidad en la UI"),
    ("unidad",              "idx_unidad_precio",         "id_proyecto, precio_lista_uf",                      "Normal",         "Ordenamiento por precio"),
    ("unidad",              "idx_unidad_condicion",      "id_proyecto, tipo_unidad, programa",                "Normal",         "JOIN con condicion_comercial"),
    ("condicion_comercial", "uq_condicion",              "id_proyecto, tipo_unidad, programa",                "Único",          "Clave natural: una condición por proyecto + tipo + programa"),
    ("condicion_comercial", "idx_condicion_proyecto",    "id_proyecto, activo",                               "Normal",         "Filtro por proyecto"),
    ("condicion_comercial", "idx_condicion_join",        "id_proyecto, tipo_unidad, programa  WHERE activo=1","Parcial activos","JOIN del cotizador solo con condiciones activas"),
    ("bien_conjunto",       "idx_bc_principal",          "id_unidad_principal",                               "Normal",         "Lookup de bienes del departamento"),
    ("bien_conjunto",       "idx_bc_asociada",           "id_unidad_asociada",                                "Normal",         "Lookup del departamento al que pertenece un bien"),
    ("uf_valor",            "idx_uf_fecha_desc",         "fecha DESC",                                        "Normal",         "Lookup eficiente del valor UF más reciente (MAX fecha)"),
    ("parametro_cotizador", "idx_parametro_categoria",   "categoria, activo",                                 "Normal",         "Carga de dropdowns de la UI"),
    ("cotizacion",          "idx_cotizacion_unidad",     "id_unidad",                                         "Normal",         "Historial de cotizaciones por unidad"),
    ("cotizacion",          "idx_cotizacion_broker",     "nombre_broker, fecha_generacion DESC",              "Normal",         "Historial de cotizaciones por broker"),
    ("cotizacion",          "idx_cotizacion_fecha",      "fecha_generacion DESC",                             "Normal",         "Historial cronológico"),
]

CHECKS = [
    # (tabla, constraint, regla)
    ("inmobiliaria",        "chk_inmobiliaria_activo",    "activo IN (0, 1)"),
    ("proyecto",            "chk_proyecto_tipo_entrega",  "tipo_entrega IN ('Entrega Inmediata', 'Entrega Futura')"),
    ("proyecto",            "chk_proyecto_activo",        "activo IN (0, 1)"),
    ("unidad",              "chk_unidad_tipo",            "tipo_unidad IN ('Departamento', 'Estacionamiento', 'Bodega', 'Local Comercial', 'Local comercial', 'Local', 'Estacionamiento Moto', 'Estacionamiento Comercial', 'Estacionamiento local')"),
    ("unidad",              "chk_unidad_estado",          "estado_stock IN ('Disponible', 'Arrendado', 'En Recolocación', 'Reservado')"),
    ("unidad",              "chk_unidad_precio",          "precio_lista_uf > 0"),
    ("unidad",              "chk_unidad_sup",             "superficie_total_m2 >= 0"),
    ("condicion_comercial", "chk_condicion_reserva",      "reserva_clp >= 0"),
    ("condicion_comercial", "chk_condicion_descuento",    "descuento >= 0 AND descuento <= 1"),
    ("condicion_comercial", "chk_condicion_bono_pie",     "bono_pie >= 0 AND bono_pie <= 1"),
    ("condicion_comercial", "chk_condicion_cuotas",       "cuotas_pie >= 0"),
    ("condicion_comercial", "chk_condicion_pie_constr",   "pie_periodo_construccion >= 0 AND pie_periodo_construccion <= 1"),
    ("condicion_comercial", "chk_condicion_cuoton",       "cuoton >= 0 AND cuoton <= 1"),
    ("condicion_comercial", "chk_condicion_cred_dir",     "pie_credito_directo >= 0 AND pie_credito_directo <= 1"),
    ("condicion_comercial", "chk_condicion_activo",       "activo IN (0, 1)"),
    ("bien_conjunto",       "chk_bc_distintos",           "id_unidad_principal != id_unidad_asociada"),
    ("uf_valor",            "chk_uf_valor",               "valor_uf > 0"),
    ("parametro_cotizador", "chk_param_categoria",        "categoria IN ('CAE', 'PIE_PCT', 'PLAZO', 'CONSTANTE')"),
    ("parametro_cotizador", "chk_param_default",          "es_default IN (0, 1)"),
    ("parametro_cotizador", "chk_param_activo",           "activo IN (0, 1)"),
]

PARAMS_DATA = [
    # (categoria, clave, valor, etiqueta, es_default)
    ("CAE",       "cae_035",    0.035, "3.5%",    0),
    ("CAE",       "cae_040",    0.040, "4.0%",    1),
    ("CAE",       "cae_045",    0.045, "4.5%",    1),
    ("CAE",       "cae_050",    0.050, "5.0%",    1),
    ("PIE_PCT",   "pie_000",    0.00,  "0%",      0),
    ("PIE_PCT",   "pie_005",    0.05,  "5%",      0),
    ("PIE_PCT",   "pie_010",    0.10,  "10%",     1),
    ("PIE_PCT",   "pie_015",    0.15,  "15%",     0),
    ("PIE_PCT",   "pie_020",    0.20,  "20%",     0),
    ("PIE_PCT",   "pie_025",    0.25,  "25%",     0),
    ("PIE_PCT",   "pie_030",    0.30,  "30%",     0),
    ("PIE_PCT",   "pie_035",    0.35,  "35%",     0),
    ("PIE_PCT",   "pie_040",    0.40,  "40%",     0),
    ("PLAZO",     "plazo_20",   20,    "20 años", 0),
    ("PLAZO",     "plazo_25",   25,    "25 años", 0),
    ("PLAZO",     "plazo_30",   30,    "30 años", 1),
    ("CONSTANTE", "UPFRONT_PCT",          0.02, "Upfront a la Promesa (%)",         0),
    ("CONSTANTE", "APORTE_INMOB_PCT",     0.10, "Aporte Inmobiliaria (%)",          0),
    ("CONSTANTE", "MESES_ARRIENDO_ANIO",  11,   "Meses de arriendo por año",        0),
    ("CONSTANTE", "HAIRCUT_VENTA",        0.95, "Factor precio venta año 5",        0),
    ("CONSTANTE", "FACTOR_LTV",           0.67, "Factor LTV amortización 60 meses", 0),
    ("CONSTANTE", "PLUSVALIA_DEFAULT",    0.02, "Plusvalía anual estimada default",  0),
]

# ── ROL → color ────────────────────────────────────────────────────────────────
ROL_BG = {"PK": C_PK_BG, "FK": C_FK_BG, "UK": C_UK_BG, "UK²": C_UK_BG,
           "—": None}

# ══════════════════════════════════════════════════════════════════════════════
# CREAR WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 1: RESUMEN
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("RESUMEN")
set_col_widths(ws, [5, 30, 14, 10, 28, 42, 38])
ws.row_dimensions[1].height = 30

# Título principal
ws.merge_cells("A1:G1")
c = ws.cell(1, 1, "MODELO DE DATOS — COTIZADOR WEB MERCADO PRIMARIO")
c.font      = Font(name="Calibri", bold=True, color=C_TITLE_FG, size=14)
c.fill      = fill(C_TITLE_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[2].height = 8
ws.merge_cells("A2:G2")
ws.cell(2,1).fill = fill("DDDDDD")

# Cabecera tabla
headers = ["#", "TABLA", "MOTOR", "FILAS INICIALES",
           "FUENTE DE DATOS", "DESCRIPCIÓN", "MANTENEDOR ADMIN"]
for col, h in enumerate(headers, 1):
    fmt_header(ws, 3, col, h, size=9)
ws.row_dimensions[3].height = 22

for i, t in enumerate(TABLES, 1):
    row = i + 3
    bg_tbl, _ = TABLE_COLORS.get(t["name"], (C_ODD, "222222"))
    bg_row = bg_tbl if i % 2 == 0 else C_EVEN
    ws.row_dimensions[row].height = 32
    fmt_cell(ws, row, 1, i,             bg=bg_row, bold=True, halign="center")
    fmt_cell(ws, row, 2, t["name"],     bg=bg_tbl, bold=True)
    fmt_cell(ws, row, 3, "SQLite",      bg=bg_row, halign="center")
    fmt_cell(ws, row, 4, f'{t["rows"]:,}' if t["rows"] else "—",
                                        bg=bg_row, halign="center")
    fmt_cell(ws, row, 5, t["fuente"],   bg=bg_row)
    fmt_cell(ws, row, 6, t["desc"],     bg=bg_row)
    fmt_cell(ws, row, 7, t["mantenedor"], bg=bg_row)

# Leyenda de roles
leg_row = len(TABLES) + 6
ws.merge_cells(start_row=leg_row, start_column=1,
               end_row=leg_row,   end_column=7)
c = ws.cell(leg_row, 1, "  LEYENDA DE ROLES DE COLUMNA")
c.font      = Font(name="Calibri", bold=True, color="444444", size=9)
c.fill      = fill(C_SECTION_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

legend = [
    ("PK", C_PK_BG, "Primary Key — clave primaria surrogate (autoincrement)"),
    ("FK", C_FK_BG, "Foreign Key — referencia a otra tabla"),
    ("UK", C_UK_BG, "Unique Key — clave alternativa o compuesta"),
    ("—",  C_EVEN,  "Columna de atributo — sin rol especial"),
]
for j, (rol, bg, desc) in enumerate(legend):
    r = leg_row + 1 + j
    fmt_cell(ws, r, 1, "",   bg=C_EVEN)
    fmt_cell(ws, r, 2, rol,  bg=bg, bold=True, halign="center")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    fmt_cell(ws, r, 3, desc, bg=C_EVEN)

ws.freeze_panes = "A4"

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 2: COLUMNAS
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("COLUMNAS")
set_col_widths(ws2, [5, 26, 28, 12, 6, 6, 18, 62])
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A1:H1")
c = ws2.cell(1, 1, "COLUMNAS POR TABLA — TIPOS, ROLES Y CONSTRAINTS")
c.font      = Font(name="Calibri", bold=True, color=C_TITLE_FG, size=13)
c.fill      = fill(C_TITLE_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

COL_HEADERS = ["#", "TABLA", "COLUMNA", "TIPO", "ROL", "NULL",
               "DEFAULT", "DESCRIPCIÓN / CONSTRAINT"]
for col, h in enumerate(COL_HEADERS, 1):
    fmt_header(ws2, 2, col, h, size=9)
ws2.row_dimensions[2].height = 22

current_row = 3
seq = 0
for t in TABLES:
    # Banda de tabla
    fmt_table_label(ws2, current_row, 1, 8, t["name"])
    current_row += 1

    for col_def in t["columns"]:
        seq += 1
        nombre, tipo, rol, nullable, default, desc = col_def
        bg_row = ROL_BG.get(rol) or (C_ODD if seq % 2 == 0 else C_EVEN)
        bg_tbl, _ = TABLE_COLORS.get(t["name"], (C_ODD, "222222"))
        lines = max(1, len(desc) // 55 + 1)
        ws2.row_dimensions[current_row].height = max(18, lines * 14)

        fmt_cell(ws2, current_row, 1, seq,      bg=bg_tbl, bold=True,  halign="center")
        fmt_cell(ws2, current_row, 2, t["name"],bg=bg_tbl, bold=True)
        fmt_cell(ws2, current_row, 3, nombre,   bg=bg_row, bold=(rol in ("PK","FK","UK","UK²")))
        fmt_cell(ws2, current_row, 4, tipo,      bg=bg_row, halign="center")
        # ROL con color propio
        rol_bg = ROL_BG.get(rol, C_EVEN) or C_EVEN
        fmt_cell(ws2, current_row, 5, rol,       bg=rol_bg, bold=True, halign="center")
        fmt_cell(ws2, current_row, 6, nullable,  bg=bg_row, halign="center")
        fmt_cell(ws2, current_row, 7, default,   bg=bg_row)
        fmt_cell(ws2, current_row, 8, desc,      bg=bg_row)
        current_row += 1

ws2.freeze_panes = "A3"

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 3: RELACIONES
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("RELACIONES")
set_col_widths(ws3, [5, 26, 38, 26, 22, 12, 12, 12, 8, 52])
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A1:J1")
c = ws3.cell(1, 1, "RELACIONES ENTRE TABLAS — FK FÍSICAS Y JOINS LÓGICOS")
c.font      = Font(name="Calibri", bold=True, color=C_TITLE_FG, size=13)
c.fill      = fill(C_TITLE_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

REL_HEADERS = ["#", "TABLA ORIGEN", "COLUMNA(S) ORIGEN",
               "TABLA DESTINO", "COLUMNA(S) DESTINO",
               "TIPO", "ON UPDATE", "ON DELETE",
               "CARD.", "DESCRIPCIÓN"]
for col, h in enumerate(REL_HEADERS, 1):
    fmt_header(ws3, 2, col, h, size=9)
ws3.row_dimensions[2].height = 22

for i, rel in enumerate(RELATIONS, 1):
    origen, col_orig, destino, col_dest, tipo, on_upd, on_del, card, desc = rel
    row = i + 2
    bg = C_REL_BG if tipo == "FK física" else "FFF2CC"
    ws3.row_dimensions[row].height = 30
    fmt_cell(ws3, row, 1,  i,        bg=bg, bold=True, halign="center")
    fmt_cell(ws3, row, 2,  origen,   bg=bg, bold=True)
    fmt_cell(ws3, row, 3,  col_orig, bg=bg)
    fmt_cell(ws3, row, 4,  destino,  bg=bg, bold=True)
    fmt_cell(ws3, row, 5,  col_dest, bg=bg)
    fmt_cell(ws3, row, 6,  tipo,     bg=bg, bold=(tipo=="FK física"), halign="center")
    fmt_cell(ws3, row, 7,  on_upd,   bg=bg, halign="center")
    fmt_cell(ws3, row, 8,  on_del,   bg=bg, halign="center")
    fmt_cell(ws3, row, 9,  card,     bg=bg, halign="center", bold=True)
    fmt_cell(ws3, row, 10, desc,     bg=bg)

ws3.freeze_panes = "A3"

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 4: ÍNDICES
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("INDICES")
set_col_widths(ws4, [5, 26, 36, 52, 32, 52])
ws4.row_dimensions[1].height = 30

ws4.merge_cells("A1:F1")
c = ws4.cell(1, 1, "ÍNDICES — OPTIMIZACIÓN DE CONSULTAS")
c.font      = Font(name="Calibri", bold=True, color=C_TITLE_FG, size=13)
c.fill      = fill(C_TITLE_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

IDX_HEADERS = ["#", "TABLA", "NOMBRE ÍNDICE", "COLUMNAS", "TIPO", "PROPÓSITO"]
for col, h in enumerate(IDX_HEADERS, 1):
    fmt_header(ws4, 2, col, h, size=9)
ws4.row_dimensions[2].height = 22

prev_table = None
for i, idx in enumerate(INDEXES, 1):
    tabla, nombre, cols, tipo, prop = idx
    row = i + 2
    bg_tbl, _ = TABLE_COLORS.get(tabla, (C_IDX_BG, "222222"))
    bg = bg_tbl if i % 2 == 0 else C_EVEN
    ws4.row_dimensions[row].height = 30
    fmt_cell(ws4, row, 1, i,       bg=bg, bold=True, halign="center")
    fmt_cell(ws4, row, 2, tabla,   bg=bg_tbl, bold=True)
    fmt_cell(ws4, row, 3, nombre,  bg=bg)
    fmt_cell(ws4, row, 4, cols,    bg=bg)
    fmt_cell(ws4, row, 5, tipo,    bg=bg, halign="center",
             bold=("Único" in tipo))
    fmt_cell(ws4, row, 6, prop,    bg=bg)

ws4.freeze_panes = "A3"

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 5: CONSTRAINTS CHECK
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("CONSTRAINTS")
set_col_widths(ws5, [5, 26, 40, 65])
ws5.row_dimensions[1].height = 30

ws5.merge_cells("A1:D1")
c = ws5.cell(1, 1, "CHECK CONSTRAINTS — VALIDACIONES A NIVEL DE BASE DE DATOS")
c.font      = Font(name="Calibri", bold=True, color=C_TITLE_FG, size=13)
c.fill      = fill(C_TITLE_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

CHK_HEADERS = ["#", "TABLA", "NOMBRE CONSTRAINT", "REGLA"]
for col, h in enumerate(CHK_HEADERS, 1):
    fmt_header(ws5, 2, col, h, size=9)
ws5.row_dimensions[2].height = 22

for i, chk in enumerate(CHECKS, 1):
    tabla, nombre, regla = chk
    row = i + 2
    bg_tbl, _ = TABLE_COLORS.get(tabla, (C_CHK_BG, "222222"))
    bg = C_CHK_BG if i % 2 == 0 else "FAF0FF"
    ws5.row_dimensions[row].height = 28
    fmt_cell(ws5, row, 1, i,       bg=bg, bold=True, halign="center")
    fmt_cell(ws5, row, 2, tabla,   bg=bg_tbl, bold=True)
    fmt_cell(ws5, row, 3, nombre,  bg=bg)
    fmt_cell(ws5, row, 4, regla,   bg=bg)

ws5.freeze_panes = "A3"

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 6: DATOS SEMILLA — parametro_cotizador
# ─────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("DATOS_SEMILLA")
set_col_widths(ws6, [5, 16, 22, 14, 28, 12])
ws6.row_dimensions[1].height = 30

ws6.merge_cells("A1:F1")
c = ws6.cell(1, 1, "DATOS SEMILLA — parametro_cotizador (22 registros iniciales)")
c.font      = Font(name="Calibri", bold=True, color=C_TITLE_FG, size=13)
c.fill      = fill(C_TITLE_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

SED_HEADERS = ["#", "CATEGORIA", "CLAVE", "VALOR", "ETIQUETA", "ES DEFAULT"]
for col, h in enumerate(SED_HEADERS, 1):
    fmt_header(ws6, 2, col, h, size=9)
ws6.row_dimensions[2].height = 22

CAT_BG = {
    "CAE":       "DDEBF7",
    "PIE_PCT":   "E2EFDA",
    "PLAZO":     "FFF2CC",
    "CONSTANTE": "F4CCFF",
}
for i, row_data in enumerate(PARAMS_DATA, 1):
    cat, clave, valor, etiqueta, es_def = row_data
    row = i + 2
    bg = CAT_BG.get(cat, C_EVEN)
    ws6.row_dimensions[row].height = 18
    fmt_cell(ws6, row, 1, i,         bg=bg, bold=True, halign="center")
    fmt_cell(ws6, row, 2, cat,       bg=bg, bold=True, halign="center")
    fmt_cell(ws6, row, 3, clave,     bg=bg)
    fmt_cell(ws6, row, 4, valor,     bg=bg, halign="right")
    fmt_cell(ws6, row, 5, etiqueta,  bg=bg)
    fmt_cell(ws6, row, 6, "SÍ" if es_def else "—",
                                     bg=bg, bold=bool(es_def), halign="center")

ws6.freeze_panes = "A3"

# ─────────────────────────────────────────────────────────────────────────────
# GUARDAR
# ─────────────────────────────────────────────────────────────────────────────
OUTPUT = "MODELO_DATOS_COTIZADOR.xlsx"
wb.save(OUTPUT)
print(f"Generado: {OUTPUT}")
print(f"  Hojas: {[s.title for s in wb.worksheets]}")
