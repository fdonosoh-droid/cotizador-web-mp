"""
upd_reglas_filtrado.py
Actualiza REGLAS_COTIZADOR.xlsx con el nuevo orden de filtrado en cascada:
  1. Comuna
  2. Entrega Aprox.   (depende de 1)
  3. Inmobiliaria     (depende de 1, 2)
  4. Proyecto         (depende de 1, 2, 3)
  5. N° Unidad        (depende de 1, 2, 3, 4)  → dispara auto-completado
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy

SRC = "REGLAS_COTIZADOR.xlsx"
DST = "REGLAS_COTIZADOR.xlsx"

wb = openpyxl.load_workbook(SRC)
ws = wb["REGLAS_COTIZADOR"]

# ── helpers ────────────────────────────────────────────────────────────────────

def copy_row_style(ws, src_row, dst_row):
    """Copia estilos de una fila a otra."""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font      = copy.copy(src.font)
            dst.fill      = copy.copy(src.fill)
            dst.alignment = copy.copy(src.alignment)
            dst.border    = copy.copy(src.border)

def set_row(ws, row_num, values):
    """Escribe una lista de valores en la fila indicada."""
    for col, val in enumerate(values, 1):
        ws.cell(row=row_num, column=col).value = val

# ── SECCIÓN 2 — header (fila 4) ───────────────────────────────────────────────
# Actualiza el título de la sección
ws.cell(row=4, column=1).value = "SECCIÓN 2 — SELECCIÓN EN CASCADA (Filtros interdependientes)"

# ── REGLA 2.1 — Selección de Comuna (nuevo primer filtro) ─────────────────────
set_row(ws, 5, [
    "2.1",
    "2. Selección en cascada",
    "Selección de Comuna",
    "Primer filtro de la cascada. El usuario elige la comuna donde busca la propiedad. "
    "Fuente: campo COMUNA de la tabla proyecto/stock.",

    "Valores: lista de comunas únicas activas en el stock.\n"
    "Ejemplo: 'La Florida', 'Pudahuel', 'Quilicura', ...\n"
    "No hay prerequisito — es el punto de entrada del cotizador.",

    "comunas = SELECT DISTINCT comuna\n"
    "          FROM proyecto\n"
    "          WHERE activo = 1\n"
    "          ORDER BY comuna ASC\n"
    "-> Dropdown de selección obligatoria",

    "Campo obligatorio. Sin selección de comuna no se habilita ningún filtro siguiente. "
    "Normalizar mayúsculas/minúsculas al cargar el stock."
])

# ── REGLA 2.2 — Entrega Aprox. (depende de Comuna) ────────────────────────────
set_row(ws, 6, [
    "2.2",
    "2. Selección en cascada",
    "Entrega Aprox.",
    "Segundo filtro. Muestra solo los tipos de entrega disponibles en la comuna seleccionada. "
    "Fuente: campo TIPO_ENTREGA de proyecto.",

    "Valores posibles: 'Entrega Inmediata' / 'Entrega Futura'\n"
    "Se filtra por la comuna elegida en 2.1.\n"
    "PREREQUISITO: 2.1 Comuna.",

    "entregas = SELECT DISTINCT tipo_entrega\n"
    "           FROM proyecto\n"
    "           WHERE comuna  = comuna_sel\n"
    "             AND activo  = 1\n"
    "           ORDER BY tipo_entrega ASC\n"
    "-> Dropdown. Se recarga al cambiar 2.1",

    "Campo obligatorio. Si solo hay un valor posible, se puede pre-seleccionar automáticamente. "
    "Bloquear si comuna_sel es NULL."
])

# ── REGLA 2.3 — Selección de Inmobiliaria (depende de Comuna + Entrega Aprox) ─
set_row(ws, 7, [
    "2.3",
    "2. Selección en cascada",
    "Selección de Inmobiliaria",
    "Tercer filtro. Lista solo las inmobiliarias que tienen proyectos en la comuna y "
    "tipo de entrega seleccionados. Fuente: tabla inmobiliaria + proyecto.",

    "Se filtra por comuna_sel (2.1) y entrega_sel (2.2).\n"
    "Ejemplo: para 'La Florida' + 'Entrega Inmediata' → solo MAESTRA.\n"
    "PREREQUISITO: 2.1 Comuna y 2.2 Entrega Aprox.",

    "inmobiliarias = SELECT DISTINCT i.nombre\n"
    "               FROM inmobiliaria i\n"
    "               JOIN proyecto p ON p.id_inmobiliaria = i.id_inmobiliaria\n"
    "               WHERE p.comuna       = comuna_sel\n"
    "                 AND p.tipo_entrega = entrega_sel\n"
    "                 AND p.activo = 1 AND i.activo = 1\n"
    "               ORDER BY i.nombre ASC\n"
    "-> Dropdown. Se recarga al cambiar 2.1 o 2.2",

    "Campo obligatorio. Se resetea al cambiar 2.1 o 2.2. "
    "Si solo hay una opción, se puede pre-seleccionar."
])

# ── REGLA 2.4 — Selección de Proyecto (depende de Comuna + Entrega + Inmobiliaria)
set_row(ws, 8, [
    "2.4",
    "2. Selección en cascada",
    "Selección de Proyecto",
    "Cuarto filtro. Muestra solo los proyectos que coinciden con los tres filtros anteriores. "
    "Fuente: tabla proyecto.",

    "Se filtra por comuna_sel (2.1), entrega_sel (2.2) e inmobiliaria_sel (2.3).\n"
    "Ejemplo: 'La Florida' + 'Entrega Inmediata' + 'MAESTRA' → 'PLAZA CERVANTES I'\n"
    "PREREQUISITO: 2.1 + 2.2 + 2.3.",

    "proyectos = SELECT id_proyecto, nombre_proyecto, periodo_entrega\n"
    "            FROM proyecto\n"
    "            WHERE comuna          = comuna_sel\n"
    "              AND tipo_entrega    = entrega_sel\n"
    "              AND id_inmobiliaria = inmobiliaria_sel\n"
    "              AND activo          = 1\n"
    "            ORDER BY nombre_proyecto ASC\n"
    "-> Dropdown. Al seleccionar, exponer periodo_entrega como campo de solo lectura",

    "Campo obligatorio. Se resetea al cambiar cualquier filtro anterior (2.1, 2.2 o 2.3). "
    "Mostrar el periodo de entrega junto al nombre del proyecto."
])

# ── SECCIÓN 3 — eliminar fila 11 (3.2 Tipo Unidad — ya no es filtro explícito)
# Antes de eliminar, verificar contenido
old_tipo_unidad = ws.cell(row=11, column=3).value
print(f"Eliminando fila 11: '{old_tipo_unidad}'")
ws.delete_rows(11)  # elimina la fila 3.2 Tipo Unidad; las siguientes suben una posición

# Después del delete, la fila que era 12 (3.3 N° Unidad) ahora es 11
# Re-numeramos: lo que era 3.3 → 3.2, 3.4 → 3.3, ... 3.11 → 3.10

renumber_map = {
    "3.3": "3.2",
    "3.4": "3.3",
    "3.5": "3.4",
    "3.6": "3.5",
    "3.7": "3.6",
    "3.8": "3.7",
    "3.9": "3.8",
    "3.10": "3.9",
    "3.11": "3.10",
}

for row in ws.iter_rows():
    cell = row[0]  # columna A = N° REGLA
    if cell.value in renumber_map:
        cell.value = renumber_map[cell.value]

# ── Actualizar regla (ahora) 3.2 — N° Unidad con nuevas dependencias ──────────
# Buscar la fila de N° Unidad por su nuevo número
for row in ws.iter_rows():
    if row[0].value == "3.2":
        r = row[0].row
        # Actualizar columnas REGLA_ACTUAL, FORMULA_CONCEPTUAL y VALIDACIONES
        ws.cell(row=r, column=3).value = "Selección de N° Unidad"
        ws.cell(row=r, column=4).value = (
            "Campo clave que dispara todos los lookups siguientes. "
            "Se muestra la lista de unidades disponibles filtrada por los cuatro "
            "criterios anteriores (2.1 → 2.4). Fuente: tabla unidad."
        )
        ws.cell(row=r, column=5).value = (
            "Lista de unidades disponibles filtradas por:\n"
            "  - comuna_sel       (de 2.1)\n"
            "  - entrega_sel      (de 2.2)\n"
            "  - inmobiliaria_sel (de 2.3)\n"
            "  - proyecto_sel     (de 2.4)\n"
            "  - estado_stock = 'Disponible'\n"
            "PREREQUISITO: 2.1 + 2.2 + 2.3 + 2.4."
        )
        ws.cell(row=r, column=6).value = (
            "unidades = SELECT id_unidad, numero_unidad, tipo_unidad, programa\n"
            "           FROM unidad\n"
            "           WHERE id_proyecto = proyecto_sel\n"
            "             AND estado_stock = 'Disponible'\n"
            "           ORDER BY numero_unidad ASC\n"
            "numero_unidad_sel = input_seleccion(unidades)\n"
            "-> Al seleccionar, se dispara auto-completado de 3.3 → 3.10"
        )
        ws.cell(row=r, column=7).value = (
            "Campo obligatorio. Solo unidades con estado_stock = 'Disponible'. "
            "Al cambiar la selección se recalculan todos los campos de solo lectura "
            "de la sección 3 (3.3 → 3.10) y todos los cálculos de secciones 4 en adelante. "
            "Si el proyecto no tiene unidades disponibles, mostrar aviso."
        )
        break

# ── Actualizar regla 3.1 — Estado de Stock para reflejar cascada ──────────────
for row in ws.iter_rows():
    if row[0].value == "3.1":
        r = row[0].row
        ws.cell(row=r, column=5).value = (
            "ESTADO STOCK = 'Disponible'  → aparece en la lista de 3.2\n"
            "ESTADO STOCK = 'Arrendado'   → NO aparece\n"
            "ESTADO STOCK = 'Reservado'   → NO aparece\n"
            "Este filtro es implícito: nunca se muestra al usuario."
        )
        ws.cell(row=r, column=6).value = (
            "unidades_cotizables = unidades.filter(\n"
            "  id_proyecto   == proyecto_sel  -- de 2.4\n"
            "  AND estado_stock == 'Disponible'\n"
            ")\n"
            "-> Regla interna, no es un campo de UI"
        )
        ws.cell(row=r, column=7).value = (
            "Regla de validación interna. Solo unidades 'Disponible' alimentan el dropdown 3.2. "
            "El filtro se aplica DENTRO del query de 3.2, no es un paso separado de UI."
        )
        break

# ── Guardar ────────────────────────────────────────────────────────────────────
wb.save(DST)
print(f"\nGuardado: {DST}")
print(f"Total filas: {ws.max_row}")

# Verificación rápida
print("\n=== Secciones 2 y 3 actualizadas ===")
for row in ws.iter_rows(min_row=4, max_row=21, values_only=True):
    num = str(row[0] or "")
    nom = str(row[2] or "")
    if num:
        print(f"  {num:12} | {nom}")
