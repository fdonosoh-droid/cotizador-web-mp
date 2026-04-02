import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook('REGLAS_PIE_Y_BONO_PIE.xlsx')
ws = wb.active

header_font   = Font(bold=True, color='FFFFFF', size=11)
header_fill   = PatternFill('solid', fgColor='1F4E79')
maestra_fill  = PatternFill('solid', fgColor='FFE699')
ingevec_fill  = PatternFill('solid', fgColor='D9EAD3')
urmeneta_fill = PatternFill('solid', fgColor='D0E4F7')
cell_font     = Font(size=10)
wrap          = Alignment(wrap_text=True, vertical='top')
center        = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin          = Side(border_style='thin', color='999999')
border        = Border(left=thin, right=thin, top=thin, bottom=thin)

UF = 39848.0

# Header col 5
c = ws.cell(row=1, column=5, value='EJEMPLO DE CALCULO (con valores reales)')
c.font = header_font
c.fill = header_fill
c.alignment = center
c.border = border

# ── MAESTRA ──────────────────────────────────────────────────────────────────
pl_depto_m = 3690.00
desc_m     = 0.10
pd_depto_m = round(pl_depto_m * (1 - desc_m), 2)
vv_m       = pd_depto_m
pie_pct_m  = 0.10
bono_pct_m = 0.10
pie_d_m    = round(pd_depto_m * pie_pct_m, 2)
pie_t_m    = pie_d_m
D33        = round(1 - pie_pct_m - bono_pct_m, 4)
tasac_m    = round(vv_m * (1 - pie_pct_m) / D33, 2)
bono_m     = round(tasac_m * bono_pct_m, 2)
cred_m     = round(tasac_m * 0.80, 2)
aporte_m   = round(tasac_m - pie_t_m - cred_m, 2)

lines_m = [
    "Datos de entrada:",
    f"  Precio lista depto : {pl_depto_m:,.2f} UF",
    f"  Descuento          : {desc_m*100:.0f}%  ->  Precio desc depto: {pd_depto_m:,.2f} UF",
    "  Sin bienes conjuntos",
    f"  %Pie: {pie_pct_m*100:.0f}%  |  %Bono Pie: {bono_pct_m*100:.0f}%",
    "",
    "PIE:",
    f"  Pie depto     = {pd_depto_m:,.2f} x {pie_pct_m*100:.0f}% = {pie_d_m:,.2f} UF",
    "  Pie conjuntos = 0 (sin bienes conjuntos)",
    f"  PIE TOTAL     = {pie_t_m:,.2f} UF  (${pie_t_m*UF:,.0f})",
    "",
    "BONO PIE (Formula D35/D36 Excel):",
    f"  D33 = 1 - {pie_pct_m*100:.0f}% - {bono_pct_m*100:.0f}% = {D33:.2f}",
    f"  Tasacion (D35) = {vv_m:,.2f} x (1-{pie_pct_m*100:.0f}%) / {D33:.2f} = {tasac_m:,.2f} UF",
    f"  Bono Pie (D36) = {tasac_m:,.2f} x {bono_pct_m*100:.0f}% = {bono_m:,.2f} UF",
    "",
    "CREDITO HIPOTECARIO:",
    f"  Tasacion Banco = {tasac_m:,.2f} UF  (${tasac_m*UF:,.0f})",
    f"  Credito Hip.   = {tasac_m:,.2f} x 80% = {cred_m:,.2f} UF  (${cred_m*UF:,.0f})",
    f"  Aporte Inmob.  = {tasac_m:,.2f} - {pie_t_m:,.2f} - {cred_m:,.2f} = {aporte_m:,.2f} UF  (${aporte_m*UF:,.0f})",
    f"  % Aporte real  = {aporte_m/tasac_m*100:.1f}% de tasacion",
]
c = ws.cell(row=2, column=5, value="\n".join(lines_m))
c.fill = maestra_fill; c.alignment = wrap; c.border = border; c.font = cell_font

# ── INGEVEC ───────────────────────────────────────────────────────────────────
pl_depto_i = 4253.00
pl_estac_i = 390.00
pl_bodeg_i = 205.00
pd_depto_i = pl_depto_i
vv_i       = pd_depto_i + pl_estac_i + pl_bodeg_i
pie_pct_i  = 0.0
bono_pct_i = 0.20
pie_d_i    = round(pd_depto_i * pie_pct_i, 2)
pie_c_i    = round((pl_estac_i + pl_bodeg_i) * 0.20, 2)
pie_t_i    = round(pie_d_i + pie_c_i, 2)
bono_i     = round(pl_depto_i * bono_pct_i, 2)
tasac_i    = round(vv_i + bono_i, 2)
cred_i     = round(vv_i - pie_t_i - bono_i, 2)

lines_i = [
    "Datos de entrada:",
    f"  Precio lista depto : {pl_depto_i:,.2f} UF",
    f"  Precio lista estac : {pl_estac_i:,.2f} UF",
    f"  Precio lista bodega: {pl_bodeg_i:,.2f} UF",
    f"  Sin descuento  ->  Valor de Venta: {vv_i:,.2f} UF",
    f"  %Pie depto: {pie_pct_i*100:.0f}%  |  %Bono Pie: {bono_pct_i*100:.0f}%",
    "",
    "PIE:",
    f"  Pie depto     = {pd_depto_i:,.2f} x {pie_pct_i*100:.0f}% = {pie_d_i:,.2f} UF",
    f"  Pie conjuntos = ({pl_estac_i:,.2f} + {pl_bodeg_i:,.2f}) x 20% = {pie_c_i:,.2f} UF",
    f"  PIE TOTAL     = {pie_d_i:,.2f} + {pie_c_i:,.2f} = {pie_t_i:,.2f} UF  (${pie_t_i*UF:,.0f})",
    "",
    "BONO PIE (base: precio lista DEPTO):",
    f"  Bono Pie  = {pl_depto_i:,.2f} x {bono_pct_i*100:.0f}% = {bono_i:,.2f} UF  (${bono_i*UF:,.0f})",
    f"  Tasacion  = {vv_i:,.2f} + {bono_i:,.2f} = {tasac_i:,.2f} UF  (${tasac_i*UF:,.0f})",
    "",
    "CREDITO HIPOTECARIO:",
    f"  Tasacion Banco = {tasac_i:,.2f} UF",
    f"  Credito Hip.   = {vv_i:,.2f} - {pie_t_i:,.2f} - {bono_i:,.2f} = {cred_i:,.2f} UF  (${cred_i*UF:,.0f})",
    f"  Aporte Inmob.  = {bono_i:,.2f} UF  ({bono_pct_i*100:.0f}% condiciones comerciales)",
]
c = ws.cell(row=3, column=5, value="\n".join(lines_i))
c.fill = ingevec_fill; c.alignment = wrap; c.border = border; c.font = cell_font

# ── URMENETA ──────────────────────────────────────────────────────────────────
pl_depto_u = 4253.00
pl_estac_u = 390.00
pl_bodeg_u = 205.00
pl_total_u = pl_depto_u + pl_estac_u + pl_bodeg_u
pd_depto_u = pl_depto_u
vv_u       = pd_depto_u + pl_estac_u + pl_bodeg_u
pie_pct_u  = 0.05
bono_pct_u = 0.15
pie_d_u    = round(pd_depto_u * pie_pct_u, 2)
pie_c_u    = round((pl_estac_u + pl_bodeg_u) * 0.20, 2)
pie_t_u    = round(pie_d_u + pie_c_u, 2)
bono_u     = round(pl_total_u * bono_pct_u, 2)
tasac_u    = round(vv_u + bono_u, 2)
cred_u     = round(vv_u - pie_t_u - bono_u, 2)

lines_u = [
    "Datos de entrada:",
    f"  Precio lista depto : {pl_depto_u:,.2f} UF",
    f"  Precio lista estac : {pl_estac_u:,.2f} UF",
    f"  Precio lista bodega: {pl_bodeg_u:,.2f} UF",
    f"  Precio lista TOTAL : {pl_total_u:,.2f} UF",
    f"  Sin descuento  ->  Valor de Venta: {vv_u:,.2f} UF",
    f"  %Pie depto: {pie_pct_u*100:.0f}%  |  %Bono Pie: {bono_pct_u*100:.0f}%",
    "",
    "PIE:",
    f"  Pie depto     = {pd_depto_u:,.2f} x {pie_pct_u*100:.0f}% = {pie_d_u:,.2f} UF",
    f"  Pie conjuntos = ({pl_estac_u:,.2f} + {pl_bodeg_u:,.2f}) x 20% = {pie_c_u:,.2f} UF",
    f"  PIE TOTAL     = {pie_d_u:,.2f} + {pie_c_u:,.2f} = {pie_t_u:,.2f} UF  (${pie_t_u*UF:,.0f})",
    "",
    "BONO PIE (base: precio lista TOTAL todas las unidades):",
    f"  Bono Pie  = {pl_total_u:,.2f} x {bono_pct_u*100:.0f}% = {bono_u:,.2f} UF  (${bono_u*UF:,.0f})",
    f"  Tasacion  = {vv_u:,.2f} + {bono_u:,.2f} = {tasac_u:,.2f} UF  (${tasac_u*UF:,.0f})",
    "",
    "CREDITO HIPOTECARIO:",
    f"  Tasacion Banco = {tasac_u:,.2f} UF",
    f"  Credito Hip.   = {vv_u:,.2f} - {pie_t_u:,.2f} - {bono_u:,.2f} = {cred_u:,.2f} UF  (${cred_u*UF:,.0f})",
    f"  Aporte Inmob.  = {bono_u:,.2f} UF  ({bono_pct_u*100:.0f}% condiciones comerciales)",
]
c = ws.cell(row=4, column=5, value="\n".join(lines_u))
c.fill = urmeneta_fill; c.alignment = wrap; c.border = border; c.font = cell_font

ws.column_dimensions['E'].width = 65
ws.row_dimensions[2].height = 180
ws.row_dimensions[3].height = 180
ws.row_dimensions[4].height = 180

wb.save('REGLAS_PIE_Y_BONO_PIE.xlsx')
print('OK')
