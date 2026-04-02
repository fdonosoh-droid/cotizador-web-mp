import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook('REGLAS_PIE_Y_BONO_PIE.xlsx')

# Eliminar hoja si ya existe
if 'REGLAS_ACTUALES_COTIZADOR' in wb.sheetnames:
    del wb['REGLAS_ACTUALES_COTIZADOR']

ws = wb.create_sheet('REGLAS_ACTUALES_COTIZADOR')

# ── Estilos ──────────────────────────────────────────────────────────────────
thin        = Side(border_style='thin', color='AAAAAA')
border      = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top    = Alignment(wrap_text=True, vertical='top')
center_mid  = Alignment(horizontal='center', vertical='center', wrap_text=True)
center_top  = Alignment(horizontal='center', vertical='top', wrap_text=True)

def hfill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def style(cell, font=None, fill=None, align=None, border_=None):
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border_: cell.border   = border_

# Colores
C_TITLE     = '1F4E79'   # azul oscuro
C_SECTION   = '2E75B6'   # azul medio
C_COMUN     = 'E2EFDA'   # verde claro
C_MAESTRA   = 'FFE699'   # amarillo
C_INGEVEC   = 'D9EAD3'   # verde
C_URMENETA  = 'D0E4F7'   # azul claro
C_SUBHEAD   = 'D6DCE4'   # gris claro
C_WHITE     = 'FFFFFF'

F_WHITE_B   = Font(bold=True, color='FFFFFF', size=12)
F_WHITE_B11 = Font(bold=True, color='FFFFFF', size=11)
F_DARK_B    = Font(bold=True, color='1F4E79', size=11)
F_DARK_B10  = Font(bold=True, color='1F4E79', size=10)
F_NORM10    = Font(size=10)
F_BOLD10    = Font(bold=True, size=10)
F_BOLD10_W  = Font(bold=True, size=10, color='FFFFFF')

# Anchos de columna
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 40

row = 1

def merged_title(ws, row, text, col_span, fill_color, font, height=22):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=text)
    style(c, font=font, fill=hfill(fill_color), align=center_mid, border_=border)
    ws.row_dimensions[row].height = height
    return row + 1

def header_row(ws, row, cols, fill_color, font, height=18):
    for i, txt in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=txt)
        style(c, font=font, fill=hfill(fill_color), align=center_mid, border_=border)
    ws.row_dimensions[row].height = height
    return row + 1

def data_row(ws, row, cols, fills, fonts, height=15):
    for i, (txt, fill_color, font) in enumerate(zip(cols, fills, fonts), 1):
        c = ws.cell(row=row, column=i, value=txt)
        style(c, font=font, fill=hfill(fill_color), align=wrap_top, border_=border)
    ws.row_dimensions[row].height = height
    return row + 1

# ── TÍTULO PRINCIPAL ─────────────────────────────────────────────────────────
row = merged_title(ws, row, 'REGLAS ACTUALES DE CÁLCULO — COTIZADOR MERCADO PRIMARIO', 5,
                   C_TITLE, F_WHITE_B, height=28)

# ── SECCIÓN 1: REGLA COMÚN ────────────────────────────────────────────────────
row = merged_title(ws, row, 'REGLA COMÚN — APLICA A TODAS LAS INMOBILIARIAS', 5,
                   C_SECTION, F_WHITE_B11, height=20)

row = header_row(ws, row,
    ['CONCEPTO', 'FÓRMULA / REGLA', 'DETALLE', 'CONDICIÓN', 'FUENTE EN CÓDIGO'],
    C_SUBHEAD, F_DARK_B10, height=18)

comun_rows = [
    ('PIE TOTAL',
     'Pie Depto + Pie Conjuntos',
     'Pie Depto = Precio Desc. Depto x %Pie\nPie Conjuntos = Precio Lista (Estac+Bodega) x 20%',
     'Bienes conjuntos SIEMPRE 20%,\nindependiente del %Pie del depto',
     'cotizador.ts: pieTotalDeptoUF\n+ pieTotalConjuntosUF'),
    ('% PIE DEFAULT EN UI',
     'max(0, 20% - %BonoPie)',
     'Ejemplo: BonoPie=20% -> Pie=0%\n         BonoPie=15% -> Pie=5%\n         BonoPie=0%  -> Pie=10%',
     'Solo cuando existe bono pie\nen condiciones comerciales',
     'PanelCotizacion.tsx: useState\n(unidad.bonoPie > 0 ? ...)'),
    ('UPFRONT PROMESA',
     'Bloqueado en 0%\ncuando hay bono pie',
     'Si bonoPie > 0 en condiciones\ncomerciales: control deshabilitado\ny valor = 0%',
     'Aplica a todas las inmobiliarias\nque tienen bono pie',
     'PanelCotizacion.tsx:\ndisabled={unidad.bonoPie > 0}'),
    ('DETECCIÓN DE INMOBILIARIA',
     'Por nombre de alianza\n(case-insensitive)',
     'getTipoCalculoBono(alianza):\n  "maestra"  -> regla Maestra\n  "urmeneta" -> regla Urmeneta\n  resto      -> regla INGEVEC/default',
     'Comparacion con .includes()\nsobre el nombre de alianza',
     'cotizadorConfig.ts:\ngetTipoCalculoBono()\ngetLtvMaxPct()'),
]

for row_data in comun_rows:
    fills  = [C_COMUN] * 5
    fonts  = [F_BOLD10, F_NORM10, F_NORM10, F_NORM10, F_NORM10]
    row = data_row(ws, row, list(row_data), fills, fonts, height=60)

# ── SECCIÓN 2: POR INMOBILIARIA ───────────────────────────────────────────────
row += 1
row = merged_title(ws, row, 'REGLAS ESPECÍFICAS POR INMOBILIARIA', 5,
                   C_SECTION, F_WHITE_B11, height=20)

# Sub-headers
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
c = ws.cell(row=row, column=1, value='Los campos en VERDE son comunes; los diferenciadores están resaltados por color de inmobiliaria')
style(c, font=F_BOLD10, fill=hfill(C_SUBHEAD), align=center_mid, border_=border)
ws.row_dimensions[row].height = 16
row += 1

row = header_row(ws, row,
    ['CONCEPTO', 'MAESTRA', 'INGEVEC (y resto default)', 'URMENETA', 'NOTAS'],
    C_SUBHEAD, F_DARK_B10, height=18)

tabla_rows = [
    (
        'Detección',
        'alianza.includes("maestra")',
        'cualquier alias no reconocido\ncomo Maestra ni Urmeneta',
        'alianza.includes("urmeneta")',
        'getTipoCalculoBono()\nen cotizadorConfig.ts'
    ),
    (
        'tipoCalculoBono',
        "'maestra'",
        "'precio-lista-depto'",
        "'precio-lista-total'",
        'Valor pasado al motor\ncomo input.tipoCalculoBono'
    ),
    (
        'LTV máximo banco\n(ltvMaxPct)',
        '0.80  (80%)',
        '1.0  (sin límite especial)',
        '1.0  (sin límite especial)',
        'getLtvMaxPct(alianza)\nen cotizadorConfig.ts'
    ),
    (
        'Base de cálculo\nBono Pie',
        'TASACIÓN BANCO\n(% se aplica sobre tasación,\nno sobre valor venta)',
        'PRECIO LISTA DEPTO\n(solo departamento,\nexcluye bienes conjuntos)',
        'PRECIO LISTA TOTAL\n(depto + estac + bodega)',
        'Diferencia clave\nentre inmobiliarias'
    ),
    (
        'Fórmula Bono Pie (UF)',
        'D36: Tasación × %BonoPie\n(Tasación se despeja primero en D35)',
        'PrecioListaDepto × %BonoPie',
        'PrecioListaTotal × %BonoPie',
        ''
    ),
    (
        'Fórmula Tasación Banco',
        'D35: ValorVenta × (1-pie)\n     / (1 - pie - bono%)\n[Excel Calculadora BP+Mutuo]',
        'ValorVenta + BonoPieUF',
        'ValorVenta + BonoPieUF',
        'Maestra: fórmula\ninversa (tasación despejada)'
    ),
    (
        'Fórmula\nCrédito Hipotecario',
        'Tasación × 80%\n(LTV máximo especial)',
        'ValorVenta - PieTotal - AporteInmob',
        'ValorVenta - PieTotal - AporteInmob',
        'Maestra tiene regla\nexclusiva de LTV 80%'
    ),
    (
        'Aporte Inmobiliaria (UF)',
        'Tasación - PieTotal - CréditoHip\n(residual, absorbe diferencia LTV)',
        'BonoPieUF calculado',
        'BonoPieUF calculado',
        ''
    ),
    (
        '% display Aporte\n(en tabla cotización)',
        'Calculado real:\nsaldoAporte / Tasación\n(puede diferir del % nominal)',
        '%BonoPie de condiciones\ncomerciales (nominal)',
        '%BonoPie de condiciones\ncomerciales (nominal)',
        'Solo Maestra muestra\n% real calculado'
    ),
]

fill_map = {
    0: [C_SUBHEAD, C_MAESTRA,  C_INGEVEC,  C_URMENETA, C_WHITE],
    1: [C_SUBHEAD, C_MAESTRA,  C_INGEVEC,  C_URMENETA, C_WHITE],
    2: [C_SUBHEAD, C_MAESTRA,  C_COMUN,    C_COMUN,    C_WHITE],
    3: [C_SUBHEAD, C_MAESTRA,  C_INGEVEC,  C_URMENETA, C_WHITE],
    4: [C_SUBHEAD, C_MAESTRA,  C_INGEVEC,  C_URMENETA, C_WHITE],
    5: [C_SUBHEAD, C_MAESTRA,  C_COMUN,    C_COMUN,    C_WHITE],
    6: [C_SUBHEAD, C_MAESTRA,  C_COMUN,    C_COMUN,    C_WHITE],
    7: [C_SUBHEAD, C_MAESTRA,  C_COMUN,    C_COMUN,    C_WHITE],
    8: [C_SUBHEAD, C_MAESTRA,  C_INGEVEC,  C_URMENETA, C_WHITE],
}

for idx, row_data in enumerate(tabla_rows):
    fills = fill_map.get(idx, [C_WHITE]*5)
    fonts = [F_BOLD10] + [F_NORM10]*4
    row = data_row(ws, row, list(row_data), fills, fonts, height=65)

# ── SECCIÓN 3: AGREGAR NUEVA INMOBILIARIA ─────────────────────────────────────
row += 1
row = merged_title(ws, row, 'COMO AGREGAR UNA NUEVA INMOBILIARIA CON REGLA ESPECIAL', 5,
                   C_SECTION, F_WHITE_B11, height=20)

instrucciones = [
    ('Paso 1 — Definir tipo\nde cálculo',
     "Editar getTipoCalculoBono()\nen lib/config/cotizadorConfig.ts",
     "Agregar:\n  if (norm.includes('nombre_alianza'))\n    return 'precio-lista-depto' | 'precio-lista-total' | 'maestra'",
     '',
     'cotizadorConfig.ts\nlineas 85-90'),
    ('Paso 2 — Definir LTV\n(si aplica)',
     "Editar getLtvMaxPct()\nen lib/config/cotizadorConfig.ts",
     "Si el banco aplica LTV especial:\n  return alianza.includes('x') ? LTV_X : 1.0",
     'Solo si el LTV difiere del estándar',
     'cotizadorConfig.ts\nlineas 72-74'),
    ('Paso 3 — Lógica de\ncálculo nueva (si aplica)',
     "Agregar bloque en\nlib/calculators/cotizador.ts",
     "En el bloque if/else de\ntipoCalculoBono:\n  } else if (tipoCalculoBono === 'nueva') {",
     'Solo si la fórmula difiere\nde las 3 existentes',
     'cotizador.ts\nlineas 220-250'),
    ('Paso 4 — Documentar',
     'Actualizar REGLAS_PIE_Y_BONO_PIE.xlsx\ny MAESTRO_DESARROLLO_COTIZADOR.md',
     '',
     '',
     ''),
]

row = header_row(ws, row,
    ['PASO', 'ACCIÓN', 'DETALLE', 'CONDICIÓN', 'ARCHIVO / LINEA'],
    C_SUBHEAD, F_DARK_B10, height=18)

for row_data in instrucciones:
    fills = [C_SUBHEAD, C_COMUN, C_WHITE, C_WHITE, C_WHITE]
    fonts = [F_BOLD10] + [F_NORM10]*4
    row = data_row(ws, row, list(row_data), fills, fonts, height=60)

# Freeze header
ws.freeze_panes = 'B4'

wb.save('REGLAS_PIE_Y_BONO_PIE.xlsx')
print('OK')
